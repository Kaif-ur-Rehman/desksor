"""
desksor/filesystem.py

Direct file system access for AI agents.
Read, write, search files and directories.
"""

import os
import fnmatch
import time
import logging
import json
import csv
from pathlib import Path
from typing import Optional

logger = logging.getLogger("desksor.filesystem")

# Maximum file size to read (10 MB)
MAX_READ_BYTES = 10 * 1024 * 1024

TEXT_EXTENSIONS = {
    ".txt", ".py", ".js", ".ts", ".jsx", ".tsx",
    ".json", ".csv", ".md", ".html", ".htm",
    ".xml", ".yaml", ".yml", ".toml", ".ini",
    ".cfg", ".conf", ".sh", ".bat", ".ps1",
    ".css", ".scss", ".less", ".sql", ".log",
    ".r", ".rb", ".go", ".rs", ".java", ".c",
    ".cpp", ".h", ".hpp", ".cs", ".php",
}


class FileSystem:
    """
    Provides AI agents with direct, safe file system access.
    All methods return {success, data, error, time_ms}.
    """

    def list_directory(self, path: str) -> dict:
        """
        List all files and folders in a directory.

        Returns: {success, data: {path, items: [{name, type, size, modified, full_path}]}, error, time_ms}
        """
        t0 = time.perf_counter()
        try:
            p = Path(path)
            if not p.exists():
                return self._err(
                    f"Path '{path}' does not exist. "
                    f"Check the path and try again.",
                    t0,
                )
            if not p.is_dir():
                return self._err(f"'{path}' is a file, not a directory.", t0)

            items = []
            for entry in sorted(p.iterdir(), key=lambda e: (e.is_file(), e.name.lower())):
                try:
                    stat = entry.stat()
                    items.append(
                        {
                            "name": entry.name,
                            "type": "file" if entry.is_file() else "directory",
                            "size": stat.st_size if entry.is_file() else None,
                            "size_human": self._human_size(stat.st_size) if entry.is_file() else None,
                            "modified": stat.st_mtime,
                            "modified_human": time.strftime(
                                "%Y-%m-%d %H:%M:%S", time.localtime(stat.st_mtime)
                            ),
                            "full_path": str(entry),
                            "extension": entry.suffix.lower() if entry.is_file() else None,
                        }
                    )
                except PermissionError:
                    items.append(
                        {
                            "name": entry.name,
                            "type": "unknown",
                            "size": None,
                            "modified": None,
                            "full_path": str(entry),
                            "error": "Permission denied",
                        }
                    )

            return self._ok(
                {"path": str(p.resolve()), "items": items, "count": len(items)},
                t0,
            )
        except PermissionError:
            return self._err(f"Permission denied reading '{path}'.", t0)
        except Exception as exc:
            logger.error("list_directory failed: %s", exc)
            return self._err(str(exc), t0)

    def read_file(self, path: str) -> dict:
        """
        Read the content of a file.
        Supports: txt, py, js, json, csv, md, html, xml, xlsx, pdf and all text files.

        Returns: {success, data: {path, content, size, extension}, error, time_ms}
        """
        t0 = time.perf_counter()
        try:
            p = Path(path)
            if not p.exists():
                return self._err(
                    f"File '{path}' does not exist. "
                    f"Use find_files() to search for the file.",
                    t0,
                )
            if not p.is_file():
                return self._err(f"'{path}' is a directory. Use list_directory() instead.", t0)

            ext = p.suffix.lower()
            stat = p.stat()
            size = stat.st_size

            if size > MAX_READ_BYTES:
                return self._err(
                    f"File '{p.name}' is {self._human_size(size)} — too large to read in full (limit 10MB). "
                    f"Use search_content() to find specific text within it.",
                    t0,
                )

            content = None

            # Excel files
            if ext in (".xlsx", ".xls", ".xlsm"):
                content = self._read_excel(p)

            # PDF files
            elif ext == ".pdf":
                content = self._read_pdf(p)

            # CSV — parse and format nicely
            elif ext == ".csv":
                content = self._read_csv(p)

            # JSON — pretty-print
            elif ext == ".json":
                content = self._read_json(p)

            # All other text files
            else:
                try:
                    content = p.read_text(encoding="utf-8", errors="replace")
                except Exception:
                    content = p.read_bytes().decode("latin-1", errors="replace")

            return self._ok(
                {
                    "path": str(p.resolve()),
                    "name": p.name,
                    "extension": ext,
                    "size": size,
                    "size_human": self._human_size(size),
                    "content": content,
                    "lines": content.count("\n") + 1 if content else 0,
                },
                t0,
            )
        except PermissionError:
            return self._err(f"Permission denied reading '{path}'.", t0)
        except Exception as exc:
            logger.error("read_file failed: %s", exc)
            return self._err(str(exc), t0)

    def write_file(self, path: str, content: str) -> dict:
        """
        Write content to a file. Creates file and parent directories if needed.

        Returns: {success, data: {path, bytes_written}, error, time_ms}
        """
        t0 = time.perf_counter()
        try:
            p = Path(path)
            p.parent.mkdir(parents=True, exist_ok=True)

            encoded = content.encode("utf-8")
            p.write_bytes(encoded)

            return self._ok(
                {
                    "path": str(p.resolve()),
                    "bytes_written": len(encoded),
                    "created": not p.existed_before if hasattr(p, "existed_before") else None,
                },
                t0,
            )
        except PermissionError:
            return self._err(f"Permission denied writing to '{path}'.", t0)
        except Exception as exc:
            logger.error("write_file failed: %s", exc)
            return self._err(str(exc), t0)

    def find_files(self, pattern: str, search_path: Optional[str] = None) -> dict:
        """
        Find files matching a glob pattern like '*.xlsx' or 'budget*'.

        Args:
            pattern: glob pattern e.g. '*.py', 'report*', 'data.xlsx'
            search_path: root directory to search (default: C:/Users)

        Returns: {success, data: {matches: [full_path, ...], count}, error, time_ms}
        """
        t0 = time.perf_counter()
        try:
            root = Path(search_path) if search_path else Path("C:/Users")
            if not root.exists():
                root = Path.home()

            matches = []
            pattern_lower = pattern.lower()

            for dirpath, dirnames, filenames in os.walk(str(root)):
                # Skip hidden and system directories
                dirnames[:] = [
                    d for d in dirnames
                    if not d.startswith(".") and d not in ("AppData", "$Recycle.Bin", "Windows")
                ]
                for filename in filenames:
                    if fnmatch.fnmatch(filename.lower(), pattern_lower):
                        matches.append(str(Path(dirpath) / filename))
                        if len(matches) >= 500:
                            break
                if len(matches) >= 500:
                    break

            return self._ok(
                {
                    "pattern": pattern,
                    "search_path": str(root),
                    "matches": matches,
                    "count": len(matches),
                    "truncated": len(matches) >= 500,
                },
                t0,
            )
        except Exception as exc:
            logger.error("find_files failed: %s", exc)
            return self._err(str(exc), t0)

    def search_content(self, query: str, path: Optional[str] = None) -> dict:
        """
        Search file contents for a query string.

        Args:
            query: text to search for
            path: directory or file to search (default: C:/Users)

        Returns: {success, data: {matches: [{file, line_number, line}], count}, error, time_ms}
        """
        t0 = time.perf_counter()
        try:
            root = Path(path) if path else Path("C:/Users")
            if not root.exists():
                root = Path.home()

            query_lower = query.lower()
            matches = []

            if root.is_file():
                files_to_search = [root]
            else:
                files_to_search = []
                for dirpath, dirnames, filenames in os.walk(str(root)):
                    dirnames[:] = [
                        d for d in dirnames
                        if not d.startswith(".") and d not in ("AppData", "$Recycle.Bin", "Windows")
                    ]
                    for filename in filenames:
                        fp = Path(dirpath) / filename
                        if fp.suffix.lower() in TEXT_EXTENSIONS:
                            files_to_search.append(fp)
                        if len(files_to_search) > 2000:
                            break

            for fp in files_to_search:
                try:
                    if fp.stat().st_size > MAX_READ_BYTES:
                        continue
                    text = fp.read_text(encoding="utf-8", errors="replace")
                    for i, line in enumerate(text.splitlines(), start=1):
                        if query_lower in line.lower():
                            matches.append(
                                {
                                    "file": str(fp),
                                    "line_number": i,
                                    "line": line.strip()[:300],
                                }
                            )
                            if len(matches) >= 200:
                                break
                except Exception:
                    pass
                if len(matches) >= 200:
                    break

            return self._ok(
                {
                    "query": query,
                    "search_path": str(root),
                    "matches": matches,
                    "count": len(matches),
                    "truncated": len(matches) >= 200,
                },
                t0,
            )
        except Exception as exc:
            logger.error("search_content failed: %s", exc)
            return self._err(str(exc), t0)

    # ─── Private file type readers ────────────────────────────────────────────

    def _read_excel(self, path: Path) -> str:
        """Read Excel file using openpyxl. Returns text representation."""
        try:
            import openpyxl

            wb = openpyxl.load_workbook(str(path), data_only=True)
            parts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                parts.append(f"=== Sheet: {sheet_name} ===")
                for row in ws.iter_rows():
                    row_values = []
                    for cell in row:
                        val = cell.value
                        row_values.append(str(val) if val is not None else "")
                    if any(v for v in row_values):
                        parts.append("\t".join(row_values))
            return "\n".join(parts)
        except ImportError:
            return "[Excel reading requires openpyxl: pip install openpyxl]"
        except Exception as exc:
            return f"[Could not read Excel file: {exc}]"

    def _read_pdf(self, path: Path) -> str:
        """Read PDF using pdfplumber. Returns extracted text."""
        try:
            import pdfplumber

            with pdfplumber.open(str(path)) as pdf:
                pages = []
                for i, page in enumerate(pdf.pages, start=1):
                    text = page.extract_text() or ""
                    pages.append(f"--- Page {i} ---\n{text}")
            return "\n\n".join(pages)
        except ImportError:
            return "[PDF reading requires pdfplumber: pip install pdfplumber]"
        except Exception as exc:
            return f"[Could not read PDF file: {exc}]"

    def _read_csv(self, path: Path) -> str:
        """Read CSV and format as aligned text table."""
        try:
            lines = []
            with open(str(path), newline="", encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f)
                for row in reader:
                    lines.append(" | ".join(str(c) for c in row))
            return "\n".join(lines)
        except Exception as exc:
            return f"[Could not read CSV: {exc}]"

    def _read_json(self, path: Path) -> str:
        """Read JSON and pretty-print it."""
        try:
            raw = path.read_text(encoding="utf-8", errors="replace")
            parsed = json.loads(raw)
            return json.dumps(parsed, indent=2, ensure_ascii=False)
        except Exception:
            return path.read_text(encoding="utf-8", errors="replace")

    @staticmethod
    def _human_size(size: int) -> str:
        """Convert bytes to human-readable size."""
        for unit in ("B", "KB", "MB", "GB"):
            if size < 1024:
                return f"{size:.1f} {unit}"
            size /= 1024
        return f"{size:.1f} TB"

    @staticmethod
    def _ok(data: dict, t0: float) -> dict:
        elapsed_ms = int((time.perf_counter() - t0) * 1000)
        return {"success": True, "data": data, "error": None, "time_ms": elapsed_ms}

    @staticmethod
    def _err(error: str, t0: float) -> dict:
        elapsed_ms = int((time.perf_counter() - t0) * 1000)
        return {"success": False, "data": None, "error": error, "time_ms": elapsed_ms}
